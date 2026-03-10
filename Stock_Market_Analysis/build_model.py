"""
build_model.py  —  Tactical Trading Analytics Pipeline
=======================================================
SOURCE : Tactical_Model_INSTITUTIONAL.xlsx  (Fact_Portfolio sheet)
OUTPUTS: Fact_Portfolio.csv | Fact_Prices.csv | Dim_Stock.csv | Dim_Scenario.csv
         tactical_dashboard.html  |  Tactical_Model_INSTITUTIONAL.xlsx (updated)

Usage  : python build_model.py
"""

import os, csv, math, logging, datetime, json, shutil, sys, time
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import BarChart, Reference

# ── Logging ────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler("build_model.log", encoding="utf-8"),
    ],
)
log = logging.getLogger("build_model")

# ── Config (edit paths here only) ─────────────────────────────────────────
BASE        = r"E:\Stock Market\Analysis\Prompts\MAR_WEEK_1 Plan"
SOURCE_XL     = os.path.join(BASE, "Tactical_Model_INSTITUTIONAL.xlsx")
OUT_XL        = os.path.join(BASE, "Tactical_Model_UPDATED.xlsx")
OUT_HTML      = os.path.join(BASE, "tactical_dashboard.html")
OUT_DIR       = BASE
BACKUP_DIR    = os.path.join(BASE, "backups")
STATUS_CSV    = os.path.join(OUT_DIR, "Pipeline_Status.csv")
TOTAL_CAPITAL = 100_000

# ── Required columns for schema validation
REQUIRED_COLS   = ["Stock", "Entry Price", "Quantity"]       # abort if missing
WARN_COLS       = ["Current Price", "Stop Loss", "Target 1", "Sector"]  # warn if missing

# ── Style tokens ───────────────────────────────────────────────────────────
BG_DARK, BG_CARD, BG_ALT = "0D1117", "161B22", "1C2128"
BLUE, GREEN, RED, GOLD, PURP = "58A6FF", "3FB950", "F85149", "E3B341", "BC8CFF"
WHITE, LIGHT, MUTED = "FFFFFF", "C9D1D9", "8B949E"

def _fill(c):   return PatternFill("solid", fgColor=c)
def _font(c=LIGHT, sz=10, bold=False):
    return Font(name="Calibri", color=c, size=sz, bold=bold)
def _align(h="center"):
    return Alignment(horizontal=h, vertical="center", wrap_text=False)
def _border():
    s = Side(style="thin", color="30363D")
    return Border(left=s, right=s, top=s, bottom=s)
def _hdr(ws, r, c, v, bg=BG_CARD, fg=GOLD):
    cell = ws.cell(row=r, column=c, value=v)
    cell.fill=_fill(bg); cell.font=_font(fg,10,True)
    cell.alignment=_align(); cell.border=_border()
def _dat(ws, r, c, v, bg=BG_DARK, fg=LIGHT, h="center", fmt=None):
    cell = ws.cell(row=r, column=c, value=v)
    cell.fill=_fill(bg); cell.font=_font(fg,10)
    cell.alignment=_align(h); cell.border=_border()
    if fmt: cell.number_format=fmt
    return cell

# ══════════════════════════════════════════════════════════════════════════
# 0. PRE-FLIGHT CHECKS  (Scenarios 7, 9, 10)
# ══════════════════════════════════════════════════════════════════════════
def preflight(source_path: str):
    """Run all pre-flight checks before touching any data."""
    # Check source exists
    if not os.path.exists(source_path):
        log.critical("SOURCE FILE NOT FOUND: %s", source_path)
        log.critical("  Create the Excel file first and add Fact_Portfolio data.")
        sys.exit(1)

    # Scenario 9 — detect Excel lock file (another user has it open)
    lock = os.path.join(os.path.dirname(source_path),
                        "~$" + os.path.basename(source_path))
    if os.path.exists(lock):
        log.error("EXCEL FILE IS LOCKED — another application has it open.")
        log.error("  Close Excel (or Power BI) holding %s, then retry.",
                  os.path.basename(source_path))
        log.error("  Lock file: %s", lock)
        sys.exit(1)

    log.info("Pre-flight OK — source file accessible")


def validate_schema(headers: list[str]) -> bool:
    """Validate required columns exist. Abort on critical missing, warn on optional."""
    ok = True
    for col in REQUIRED_COLS:
        found = any(col.lower() == h.lower().strip() for h in headers)
        if not found:
            log.critical("REQUIRED COLUMN MISSING: '%s'", col)
            log.critical("  Found columns: %s", headers)
            ok = False
    for col in WARN_COLS:
        found = any(col.lower() == h.lower().strip() for h in headers)
        if not found:
            log.warning("Optional column missing: '%s' — defaults will be used", col)
    if not ok:
        log.critical("Aborting — fix column names in Excel and retry.")
        sys.exit(1)
    return True


def backup_source(source_path: str):
    """Scenario 10 — create dated backup of source Excel before processing."""
    os.makedirs(BACKUP_DIR, exist_ok=True)
    today = datetime.date.today().strftime("%Y%m%d")
    base  = os.path.splitext(os.path.basename(source_path))[0]
    dest  = os.path.join(BACKUP_DIR, f"{base}_{today}.xlsx")
    if not os.path.exists(dest):          # only backup once per day
        shutil.copy2(source_path, dest)
        log.info("Backup created: %s", os.path.basename(dest))
    else:
        log.info("Backup already exists for today: %s", os.path.basename(dest))



def load_portfolio(path: str) -> list[dict]:
    """Read Fact_Portfolio sheet. Returns list of dicts keyed by header."""
    log.info("Loading source: %s", path)
    if not os.path.exists(path):
        raise FileNotFoundError(f"Source not found: {path}")
    wb = openpyxl.load_workbook(path, data_only=True)

    # Prefer 'Fact_Portfolio', fall back to first sheet
    sheet_name = "Fact_Portfolio" if "Fact_Portfolio" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Fact_Portfolio sheet is empty")

    # Auto-detect: if row[0] is a merged title (only 1 non-None value), skip it
    # and treat row[1] as the real headers
    non_none_0 = sum(1 for v in rows[0] if v is not None)
    if non_none_0 <= 2 and len(rows) > 2:
        header_row = rows[1]
        data_start = 2
    else:
        header_row = rows[0]
        data_start = 1

    headers = [str(h).strip() if h else f"Col{i}" for i, h in enumerate(header_row, 1)]
    records = []
    for row in rows[data_start:]:
        if not row[0]:          # skip blank rows
            continue
        records.append(dict(zip(headers, row)))

    log.info("  Loaded %d stocks from '%s'", len(records), sheet_name)
    validate_schema(headers)    # abort here if required columns missing
    return records


def load_scenarios(path: str):
    """Read Dim_Scenario sheet if present, else return hardcoded defaults."""
    wb = openpyxl.load_workbook(path, data_only=True)
    defaults = [
        ("Bullish", 0.30, 105685.61, 0.08,  0.06, -0.02, "Oil spike scenario"),
        ("Neutral", 0.40,  99814.19, 0.02,  0.02,  0.00, "Range-bound scenario"),
        ("Bearish", 0.30,  92964.20, -0.05, -0.03, -0.04, "Ceasefire / selloff"),
    ]
    if "Dim_Scenario" not in wb.sheetnames:
        return defaults
    ws   = wb["Dim_Scenario"]
    rows = list(ws.iter_rows(values_only=True))
    # Skip any rows where first value is non-numeric (title / header rows)
    data = [r for r in rows if r[0] and isinstance(r[0], str)
            and r[0].strip() not in ("",) and not str(r[0]).startswith("Scenario") and not str(r[0]).startswith("📊") and not str(r[0]).startswith("🎯")]
    # Better: skip rows where column 2 (probability) is not numeric
    out = []
    for r in rows:
        try:
            if r[0] and float(r[1]) <= 1.0:
                out.append(tuple(r[:7]))
        except (TypeError, ValueError, IndexError):
            continue
    return out if out else defaults


def load_prices(path: str):
    """Read Fact_Prices sheet for time-series data."""
    wb = openpyxl.load_workbook(path, data_only=True)
    if "Fact_Prices" not in wb.sheetnames:
        return []
    ws   = wb["Fact_Prices"]
    rows = list(ws.iter_rows(values_only=True))
    out  = []
    for r in rows:
        # Skip header / title rows — data rows have a numeric price in col 3
        try:
            if r[0] and float(r[2] if len(r)>2 else r[-1]) > 0:
                out.append(r)
        except (TypeError, ValueError):
            continue
    return out

# ══════════════════════════════════════════════════════════════════════════
# 2. COMPUTE METRICS  (pure functions — no side effects)
# ══════════════════════════════════════════════════════════════════════════
def _f(v, d=0.0):
    """Safe float conversion."""
    try:
        x = float(v)
        return x if math.isfinite(x) else d
    except (TypeError, ValueError):
        return d

def quant_action(pnl_pct: float) -> str:
    if pnl_pct >=  0.15: return "Reduce 30%"
    if pnl_pct >=  0.10: return "Book 40%"
    if pnl_pct >=  0.06: return "Book 30%"
    if pnl_pct <= -0.05: return "EXIT NOW"
    return "Hold"

def compute_row(rec: dict, today: str) -> dict:
    """
    Compute all derived fields for one stock.
    Auto-fills missing Target 1/2/3, Stop Loss, Risk Level,
    Strategic Role, Volatility from Entry Price when blank.
    """
    # ── Static fields
    stock    = str(rec.get("Stock","")).strip()
    sector   = str(rec.get("Sector","")).strip()
    subsect  = str(rec.get("Sub-Sector", rec.get("Sub Sector",""))).strip()
    entry    = _f(rec.get("Entry Price", rec.get("Entry",0)))
    qty      = int(_f(rec.get("Quantity", rec.get("Qty",0))))
    alloc    = _f(rec.get("Allocation ₹", rec.get("Allocation",0)))

    # ── Auto-fill targets and stop loss if blank ───────────────────
    sl  = _f(rec.get("Stop Loss",""))
    t1  = _f(rec.get("Target 1",""))
    t2  = _f(rec.get("Target 2",""))
    t3  = _f(rec.get("Target 3",""))

    if entry > 0:
        if sl  == 0.0:
            sl = round(entry * 0.95, 2)   # default SL = -5%
            log.warning("  [%s] Stop Loss blank — auto-set to -5%% of Entry (%.2f)", stock, sl)
        if t1  == 0.0:
            t1 = round(entry * 1.06, 2)   # default T1 = +6%
            log.warning("  [%s] Target 1 blank — auto-set to +6%% of Entry (%.2f)", stock, t1)
        if t2  == 0.0:
            t2 = round(entry * 1.10, 2)   # default T2 = +10%
            log.warning("  [%s] Target 2 blank — auto-set to +10%% of Entry (%.2f)", stock, t2)
        if t3  == 0.0:
            t3 = round(entry * 1.15, 2)   # default T3 = +15%
            log.warning("  [%s] Target 3 blank — auto-set to +15%% of Entry (%.2f)", stock, t3)

    # ── Auto-fill qualitative fields if blank ─────────────────────
    risk_lvl_raw = rec.get("Risk Level", "")
    role_raw     = rec.get("Strategic Role", rec.get("Role",""))
    vol_raw      = rec.get("Volatility","")

    risk_lvl = str(risk_lvl_raw).strip() if risk_lvl_raw and str(risk_lvl_raw).strip() not in ("","None","nan") else "Medium"
    role     = str(role_raw).strip()     if role_raw     and str(role_raw).strip()     not in ("","None","nan") else "Tactical Swing"
    vol      = str(vol_raw).strip()      if vol_raw      and str(vol_raw).strip()      not in ("","None","nan") else "Medium"

    # ── Dynamic fields (user updates Current Price daily)
    cur = _f(rec.get("Current Price", rec.get("Current", 0)))
    # Scenario 4 — blank/zero current price falls back to entry price
    if cur == 0.0 and entry > 0:
        log.warning("  [%s] Current Price is blank — using Entry Price (%.2f) as fallback",
                    stock, entry)
        cur = entry

    # ── Computed metrics
    inv_val  = round(entry * qty, 2)
    alloc    = inv_val                      # SYNC: Allocation is now equal to Investment Value
    cur_val  = round(cur   * qty, 2)
    pnl_inr  = round(cur_val - inv_val, 2)
    pnl_pct  = round(pnl_inr / inv_val, 6) if inv_val else 0.0
    risk_amt = round((entry - sl) * qty, 2) if sl else 0.0
    sl_exit  = round(sl * qty, 2)           # SL Exit Value = SL price × Qty
    dist_sl  = round((cur - sl) / cur, 6)   if cur and sl else 0.0
    dist_t1  = round((t1  - cur) / cur, 6)  if cur and t1 else 0.0
    rr       = round((t1 - entry) / (entry - sl), 2) if (entry - sl) > 0 else 0.0
    ev       = round(rr * risk_amt, 2)
    action   = quant_action(pnl_pct)

    return {
        "Stock": stock, "Sector": sector, "Sub-Sector": subsect,
        "Allocation Rs": alloc, "Entry Price": entry, "Quantity": qty,
        "Target 1": t1, "Target 2": t2, "Target 3": t3, "Stop Loss": sl,
        "Risk Level": risk_lvl, "Investment Value": inv_val,
        "Current Price": cur, "Current Value": cur_val,
        "PnL Rs": pnl_inr, "PnL Pct": pnl_pct,
        "Risk Amount Rs": risk_amt, "SL Exit Value": sl_exit,
        "Dist to SL Pct": dist_sl, "Dist to T1 Pct": dist_t1,
        "Reward Risk": rr, "Expected Value": ev,
        "Action": action, "Strategic Role": role, "Volatility": vol, "Date": today,
    }

def compute_kpis(rows: list[dict], total_capital: float) -> dict:
    invested  = sum(r["Investment Value"] for r in rows)
    cur_val   = sum(r["Current Value"]    for r in rows)
    pnl_inr   = round(cur_val - invested, 2)
    pnl_pct   = round(pnl_inr / invested * 100, 4) if invested else 0
    total_risk= sum(r["Risk Amount Rs"]   for r in rows)
    heat      = round(total_risk / invested * 100, 4) if invested else 0
    return {
        "Total Capital":           total_capital,
        "Total Invested":          round(invested, 2),
        "Total Current Value":     round(cur_val, 2),
        "Total PnL Rs":            pnl_inr,
        "Total PnL Pct":           pnl_pct,
        "Total Risk Rs":           round(total_risk, 2),
        "Capital After SL":        round(invested - total_risk, 2),
        "Portfolio Heat Pct":      heat,
        "Risk Status":             "HARD STOP" if heat>=6 else "Warning" if heat>=4 else "Safe",
        "Active Stocks":           len(rows),
        "Cash Reserve":            round(total_capital - invested, 2),
    }

# ══════════════════════════════════════════════════════════════════════════
# 3. EXPORT CSVs
# ══════════════════════════════════════════════════════════════════════════
FP_HEADERS = [
    "Stock","Sector","Sub-Sector","Allocation Rs","Entry Price","Quantity",
    "Target 1","Target 2","Target 3","Stop Loss","Risk Level",
    "Investment Value","Current Price","Current Value","PnL Rs","PnL Pct",
    "Risk Amount Rs","SL Exit Value","Dist to SL Pct","Dist to T1 Pct",
    "Reward Risk","Expected Value","Action","Strategic Role","Volatility","Date"
]

def export_csvs(rows: list[dict], scenarios, price_rows, out_dir: str):
    # Fact_Portfolio
    fp = os.path.join(out_dir, "Fact_Portfolio.csv")
    with open(fp, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=FP_HEADERS, extrasaction="ignore")
        w.writeheader(); w.writerows(rows)
    log.info("  CSV: Fact_Portfolio.csv  (%d rows)", len(rows))

    # Dim_Stock — derived dynamically from rows (no hardcoded dict)
    seen, dim_rows = set(), []
    for r in rows:
        if r["Stock"] not in seen:
            seen.add(r["Stock"])
            dim_rows.append({
                "Stock Name": r["Stock"], "Sector": r["Sector"],
                "Sub-Sector": r["Sub-Sector"], "Volatility": r["Volatility"],
                "Strategic Role": r["Strategic Role"],
            })
    ds = os.path.join(out_dir, "Dim_Stock.csv")
    with open(ds, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["Stock Name","Sector","Sub-Sector","Volatility","Strategic Role"])
        w.writeheader(); w.writerows(dim_rows)
    log.info("  CSV: Dim_Stock.csv  (%d stocks)", len(dim_rows))

    # Dim_Scenario
    dsc = os.path.join(out_dir, "Dim_Scenario.csv")
    with open(dsc, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["Scenario","Probability","Projected Value","Oil Chg","Gold Chg","Nifty Chg","Description"])
        for sc in scenarios: w.writerow(list(sc))
    log.info("  CSV: Dim_Scenario.csv")

    # Fact_Prices
    fpr = os.path.join(out_dir, "Fact_Prices.csv")
    with open(fpr, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["Date","Stock","Price"])
        for r in price_rows: w.writerow(list(r)[:3])
    log.info("  CSV: Fact_Prices.csv  (%d rows)", len(price_rows))

# ══════════════════════════════════════════════════════════════════════════
# 4. WRITE EXCEL (update Tactical_Model_INSTITUTIONAL.xlsx)
# ══════════════════════════════════════════════════════════════════════════
def write_excel(rows: list[dict], kpis: dict, scenarios, price_rows, path: str):
    wb = Workbook(); wb.remove(wb.active)
    _write_fact_portfolio(wb, rows)
    _write_dim_stock(wb, rows)
    _write_dim_scenario(wb, scenarios)
    _write_fact_prices(wb, price_rows)
    _write_portfolio_summary(wb, kpis)
    _write_viz_charts(wb, rows)
    wb.active = wb["Portfolio_Summary"]
    try:
        wb.save(path)
        log.info("Excel saved: %s", os.path.basename(path))
    except PermissionError:
        fallback = path.replace(".xlsx", "_temp.xlsx")
        wb.save(fallback)
        log.warning("Permission denied on %s — saved to %s instead", os.path.basename(path), os.path.basename(fallback))
        log.warning("  (Close the file in Excel, then rename/replace manually)")

def _write_fact_portfolio(wb, rows):
    ws = wb.create_sheet("Portfolio_Master")
    ws.sheet_properties.tabColor = BLUE
    ws.freeze_panes = "A2"
    for ci, h in enumerate(FP_HEADERS, 1):
        _hdr(ws, 1, ci, h)
    
    # Formula mapping: Target 1 (Col 7) = [@Entry Price] * 1.06
    # Note: Column 5 is Entry Price, Column 7 is Target 1
    
    FMT = {"Entry Price":"#,##0.00","Current Price":"#,##0.00",
           "Target 1":"#,##0.00","Target 2":"#,##0.00","Target 3":"#,##0.00",
           "Stop Loss":"#,##0.00","Investment Value":"#,##0.00",
           "Current Value":"#,##0.00","PnL Rs":"#,##0.00",
           "Allocation Rs":"#,##0.00","Risk Amount Rs":"#,##0.00",
           "SL Exit Value":"#,##0.00","Expected Value":"#,##0.00",
           "PnL Pct":"0.00%","Dist to SL Pct":"0.00%",
           "Dist to T1 Pct":"0.00%","Reward Risk":"0.00"}
    
    for ri, row in enumerate(rows, 2):
        bg = BG_DARK if ri%2==0 else BG_ALT
        for ci, h in enumerate(FP_HEADERS, 1):
            v = row.get(h,"")
            fg = LIGHT
            if h == "PnL Rs" or h == "PnL Pct":
                fg = GREEN if _f(v)>=0 else RED
            elif h == "Action":
                fg = RED if v=="EXIT NOW" else GOLD if v!="Hold" else MUTED
            
            # Scenario 11/Phase 6 — Reapply formula for Target 1
            if h == "Target 1":
                # Structured reference formula (openpyxl creates the table below)
                # Or row-based formula: =E{ri}*1.06
                v = f"=E{ri}*1.06"
            
            _dat(ws, ri, ci, v, bg, fg, fmt=FMT.get(h))
    
    ws.column_dimensions["A"].width = 26
    for ci in range(2, len(FP_HEADERS)+1):
        ws.column_dimensions[get_column_letter(ci)].width = 14

    # Convert to Excel Table for auto-expansion
    table_ref = f"A1:{get_column_letter(len(FP_HEADERS))}{len(rows)+1}"
    tab = Table(displayName="FactPortfolio", ref=table_ref)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)

def _write_dim_stock(wb, rows):
    ws = wb.create_sheet("Dim_Stock"); ws.sheet_properties.tabColor = PURP
    ws.freeze_panes = "A2"
    heads = ["Stock Name","Sector","Sub-Sector","Volatility","Strategic Role"]
    for ci,h in enumerate(heads,1): _hdr(ws,1,ci,h)
    seen = set()
    ri = 2
    for r in rows:
        if r["Stock"] in seen: continue
        seen.add(r["Stock"])
        bg = BG_DARK if ri%2==0 else BG_ALT
        vals = [r["Stock"],r["Sector"],r["Sub-Sector"],r["Volatility"],r["Strategic Role"]]
        for ci,v in enumerate(vals,1): _dat(ws,ri,ci,v,bg,h="left" if ci==1 else "center")
        ri+=1
    for ci,w in enumerate([28,14,18,14,18],1):
        ws.column_dimensions[get_column_letter(ci)].width=w

def _write_dim_scenario(wb, scenarios):
    ws = wb.create_sheet("Dim_Scenario"); ws.sheet_properties.tabColor = GOLD
    heads = ["Scenario","Probability","Projected Value","Oil Chg","Gold Chg","Nifty Chg","Description"]
    for ci,h in enumerate(heads,1): _hdr(ws,1,ci,h)
    colors = [GREEN, BLUE, RED]
    for ri, (sc, clr) in enumerate(zip(scenarios, colors), 2):
        bg = BG_DARK if ri%2==0 else BG_ALT
        for ci,v in enumerate(sc,1):
            fg = clr if ci==1 else LIGHT
            _dat(ws,ri,ci,v,bg,fg)
    for ci,w in enumerate([12,12,18,10,10,10,40],1):
        ws.column_dimensions[get_column_letter(ci)].width=w

def _write_fact_prices(wb, price_rows):
    ws = wb.create_sheet("Fact_Prices"); ws.sheet_properties.tabColor = GREEN
    ws.freeze_panes = "A2"
    for ci,h in enumerate(["Date","Stock","Price"],1): _hdr(ws,1,ci,h)
    for ri,r in enumerate(price_rows,2):
        bg = BG_DARK if ri%2==0 else BG_ALT
        vals = list(r)[:3]
        for ci,v in enumerate(vals,1):
            fmt = "#,##0.00" if ci==3 else None
            _dat(ws,ri,ci,v,bg,fmt=fmt)
    for ci,w in enumerate([14,24,14],1):
        ws.column_dimensions[get_column_letter(ci)].width=w

def _write_portfolio_summary(wb, kpis):
    ws = wb.create_sheet("Portfolio_Summary"); ws.sheet_properties.tabColor = RED
    for ci,h in enumerate(["Metric","Value","Status"],1): _hdr(ws,1,ci,h)
    for ri,(k,v) in enumerate(kpis.items(),2):
        bg = BG_DARK if ri%2==0 else BG_ALT
        _dat(ws,ri,1,k,bg,WHITE,h="left")
        fg = GREEN if "Safe" in str(v) else RED if "STOP" in str(v) or "Warning" in str(v) else LIGHT
        if "PnL" in k: fg = GREEN if _f(v)>=0 else RED
        _dat(ws,ri,2,v,bg,fg)
        status = "✅ OK" if fg==GREEN else "⚠ Warn" if fg==GOLD else "🛑" if fg==RED else "—"
        _dat(ws,ri,3,status,bg)
    for ci,w in enumerate([30,20,12],1):
        ws.column_dimensions[get_column_letter(ci)].width=w

def _write_viz_charts(wb, rows):
    ws = wb.create_sheet("Stock_Visualization_Data"); ws.sheet_properties.tabColor = GOLD
    heads = ["Stock","Entry","Current","Stop Loss","Target 1","Target 2","Target 3"]
    for ci,h in enumerate(heads,1): _hdr(ws,1,ci,h)
    for ri,r in enumerate(rows,2):
        bg = BG_DARK if ri%2==0 else BG_ALT
        
        # Scenario 11/Phase 6 — XLOOKUP Target 3 from Portfolio_Master
        # Target 3 is Column I (9) in Portfolio_Master. Stock is Column A (1).
        formula_g = f"=XLOOKUP(A{ri}, Portfolio_Master!$A:$A, Portfolio_Master!$I:$I)"
        
        vals = [r["Stock"],r["Entry Price"],r["Current Price"],
                r["Stop Loss"],r["Target 1"],r["Target 2"], formula_g]
        for ci,v in enumerate(vals,1):
            fmt = "#,##0.00" if ci>1 else None
            _dat(ws,ri,ci,v,bg,h="left" if ci==1 else "center",fmt=fmt)
    
    # Convert to Excel Table for dynamic chart ranges
    table_ref = f"A1:G{len(rows)+1}"
    tab = Table(displayName="StockVizData", ref=table_ref)
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)
    
    n = len(rows)+1
    for title,mnc,mxc,anchor in [
        ("Entry vs Current",2,3,"A4"),("Stop Loss Levels",4,4,"J4"),
        ("Target 1",5,5,"A22"),("Target 2–3",6,7,"J22")]:
        ch = BarChart(); ch.type="col"; ch.title=title; ch.style=11
        ch.width=18; ch.height=12
        ch.add_data(Reference(ws,min_col=mnc,max_col=mxc,min_row=1,max_row=n),titles_from_data=True)
        ch.set_categories(Reference(ws,min_col=1,min_row=2,max_row=n))
        ws.add_chart(ch,anchor)
    for ci,w in enumerate([26,12,12,12,12,12,12],1):
        ws.column_dimensions[get_column_letter(ci)].width=w

# ══════════════════════════════════════════════════════════════════════════
# 5. GENERATE HTML DASHBOARD  (fully dynamic from computed data)
# ══════════════════════════════════════════════════════════════════════════
def generate_html(rows: list[dict], kpis: dict, scenarios, out_path: str):
    today = rows[0]["Date"] if rows else datetime.date.today().isoformat()

    # Build JSON for JS
    stocks_js   = json.dumps(rows, default=str)
    kpis_js     = json.dumps(kpis, default=str)

    # Scenario projection data
    sc_labels = [sc[0] for sc in scenarios]
    sc_vals   = [float(sc[2]) if sc[2] else 0 for sc in scenarios]
    sc_probs  = [float(sc[1]) if sc[1] else 0 for sc in scenarios]
    invested  = kpis["Total Invested"]
    sc_js = json.dumps([{"label":l,"value":v,"prob":p}
                         for l,v,p in zip(sc_labels,sc_vals,sc_probs)])

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width,initial-scale=1"/>
<title>Tactical Portfolio Dashboard — {today}</title>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;600;700;800&family=JetBrains+Mono:wght@400;600&display=swap" rel="stylesheet"/>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<style>
:root{{--bg:#0D1117;--bg2:#161B22;--bg3:#1C2128;--b:#30363D;
  --blue:#58A6FF;--green:#3FB950;--red:#F85149;--gold:#E3B341;--purp:#BC8CFF;
  --text:#F0F6FC;--muted:#8B949E;--r:10px;}}
*{{box-sizing:border-box;margin:0;padding:0}}
body{{background:var(--bg);color:var(--text);font-family:'Inter',sans-serif;min-height:100vh}}
body::before{{content:'';position:fixed;inset:0;z-index:0;pointer-events:none;
  background-image:linear-gradient(rgba(88,166,255,.03)1px,transparent 1px),linear-gradient(90deg,rgba(88,166,255,.03)1px,transparent 1px);
  background-size:40px 40px}}
.wrap{{max-width:1500px;margin:0 auto;padding:20px 22px;position:relative;z-index:1}}
.hdr{{display:flex;align-items:center;justify-content:space-between;padding:22px 28px;
  background:var(--bg2);border:1px solid var(--b);border-radius:var(--r);margin-bottom:20px}}
.hdr h1{{font-size:20px;font-weight:800;background:linear-gradient(90deg,var(--blue),var(--purp));
  -webkit-background-clip:text;-webkit-text-fill-color:transparent}}
.hdr p{{font-size:12px;color:var(--muted);margin-top:4px;font-family:'JetBrains Mono',monospace}}
.badges{{display:flex;gap:8px;flex-wrap:wrap}}
.badge{{padding:5px 12px;border-radius:20px;font-size:11px;font-weight:700;letter-spacing:.4px}}
.b-blue{{background:rgba(88,166,255,.15);color:var(--blue);border:1px solid rgba(88,166,255,.3)}}
.b-green{{background:rgba(63,185,80,.15);color:var(--green);border:1px solid rgba(63,185,80,.3)}}
.b-gold{{background:rgba(227,179,65,.15);color:var(--gold);border:1px solid rgba(227,179,65,.3)}}
.b-red{{background:rgba(248,81,73,.2);color:var(--red);border:1px solid rgba(248,81,73,.4)}}
.kpi-row{{display:grid;grid-template-columns:repeat(auto-fit,minmax(160px,1fr));gap:14px;margin-bottom:20px}}
.kpi{{background:var(--bg2);border:1px solid var(--b);border-radius:var(--r);padding:18px;position:relative;overflow:hidden;transition:transform .2s}}
.kpi:hover{{transform:translateY(-2px)}}
.kpi::before{{content:'';position:absolute;top:0;left:0;right:0;height:2px}}
.kpi.ck-blue::before{{background:var(--blue)}} .kpi.ck-green::before{{background:var(--green)}}
.kpi.ck-red::before{{background:var(--red)}} .kpi.ck-gold::before{{background:var(--gold)}}
.kpi.ck-purp::before{{background:var(--purp)}}
.kpi-label{{font-size:10px;color:var(--muted);text-transform:uppercase;letter-spacing:.5px;margin-bottom:6px}}
.kpi-val{{font-size:20px;font-weight:800;font-family:'JetBrains Mono',monospace}}
.kpi-sub{{font-size:11px;color:var(--muted);margin-top:5px}}
.grid3{{display:grid;grid-template-columns:1fr 1fr 1fr;gap:16px;margin-bottom:16px}}
.grid2{{display:grid;grid-template-columns:2fr 1fr;gap:16px;margin-bottom:16px}}
.panel{{background:var(--bg2);border:1px solid var(--b);border-radius:var(--r);padding:20px}}
.panel-hdr{{display:flex;justify-content:space-between;align-items:center;margin-bottom:14px}}
.panel-title{{font-size:13px;font-weight:700;display:flex;align-items:center;gap:6px}}
.dot{{width:7px;height:7px;border-radius:50%;display:inline-block}}
.panel-sub{{font-size:11px;color:var(--muted)}}
.tbl-wrap{{overflow-x:auto;border-radius:6px}}
.tbl-h{{max-height:340px;overflow-y:auto}}
.tbl-h::-webkit-scrollbar{{width:4px}} .tbl-h::-webkit-scrollbar-thumb{{background:var(--b);border-radius:2px}}
table{{width:100%;border-collapse:collapse;font-size:12px}}
thead th{{background:var(--bg3);color:var(--muted);font-weight:600;padding:9px 11px;
  text-align:left;white-space:nowrap;border-bottom:1px solid var(--b);position:sticky;top:0}}
tbody tr{{border-bottom:1px solid rgba(48,54,61,.4);transition:background .15s}}
tbody tr:hover{{background:var(--bg3)}}
tbody td{{padding:10px 11px;white-space:nowrap;color:var(--muted)}}
tbody td:first-child{{color:var(--text);font-weight:600}}
.mono{{font-family:'JetBrains Mono',monospace;font-size:11px}}
.pos{{color:var(--green)!important}} .neg{{color:var(--red)!important}} .warn{{color:var(--gold)!important}}
.ab{{padding:3px 9px;border-radius:10px;font-size:10px;font-weight:700;display:inline-block}}
.ab-hold{{background:rgba(139,148,158,.1);color:var(--muted);border:1px solid rgba(139,148,158,.2)}}
.ab-book{{background:rgba(227,179,65,.12);color:var(--gold);border:1px solid rgba(227,179,65,.3)}}
.ab-reduce{{background:rgba(88,166,255,.12);color:var(--blue);border:1px solid rgba(88,166,255,.3)}}
.ab-exit{{background:rgba(248,81,73,.2);color:var(--red);border:1px solid rgba(248,81,73,.4);animation:pulse .8s ease-in-out infinite alternate}}
@keyframes pulse{{from{{opacity:1}}to{{opacity:.5}}}}
.sb-row{{display:flex;align-items:center;gap:8px;margin-bottom:8px}}
.sb-label{{font-size:12px;width:120px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}}
.sb-track{{flex:1;height:6px;background:var(--bg3);border-radius:3px;overflow:hidden}}
.sb-fill{{height:100%;border-radius:3px;transition:width .6s ease}}
.sb-pct{{font-size:11px;font-family:'JetBrains Mono',monospace;color:var(--muted);width:42px;text-align:right}}
.sc-grid{{display:grid;grid-template-columns:repeat(3,1fr);gap:14px;margin-bottom:16px}}
.sc{{background:var(--bg3);border:1px solid var(--b);border-radius:var(--r);padding:16px;text-align:center}}
.sc-name{{font-size:12px;font-weight:700;margin-bottom:6px}}
.sc-val{{font-size:20px;font-weight:900;font-family:'JetBrains Mono',monospace}}
.sc-prob{{font-size:11px;color:var(--muted);margin-top:6px}}
.ctrl{{display:flex;gap:10px;flex-wrap:wrap;margin-bottom:16px;align-items:center}}
.ctrl-panel{{background:var(--bg2);border:1px solid var(--b);border-radius:6px;padding:8px 14px;display:flex;align-items:center;gap:8px}}
.ctrl-label{{font-size:11px;color:var(--muted)}}
select{{background:var(--bg3);border:1px solid var(--b);color:var(--text);border-radius:5px;padding:6px 10px;font-size:12px;outline:none;font-family:'Inter',sans-serif}}
.footer{{text-align:center;padding:20px;color:var(--muted);font-size:11px;border-top:1px solid var(--b);margin-top:16px}}
@media(max-width:1000px){{.grid3,.grid2{{grid-template-columns:1fr}} .sc-grid{{grid-template-columns:1fr}}}}
</style>
</head>
<body>
<div class="wrap">
<div class="hdr">
  <div>
    <h1>📊 Tactical Portfolio Intelligence Dashboard</h1>
    <p id="hdrSub">Loading…</p>
  </div>
  <div class="badges">
    <span class="badge b-blue">⚡ Dynamic</span>
    <span class="badge b-green">✅ Power BI Ready</span>
    <span class="badge b-gold">🏛 Institutional</span>
    <span class="badge b-red" id="riskBadge">⚠ Risk Monitor</span>
  </div>
</div>

<div class="ctrl">
  <div class="ctrl-panel">
    <span class="ctrl-label">SECTOR</span>
    <select id="sectorFilter" onchange="filterSector(this.value)"><option value="all">All Sectors</option></select>
  </div>
  <div class="ctrl-panel" style="margin-left:auto">
    <span class="ctrl-label">Updated:</span>
    <span class="mono" id="updDate"></span>
  </div>
</div>

<div class="kpi-row" id="kpiRow"></div>

<div class="grid3">
  <div class="panel"><div class="panel-hdr"><div class="panel-title"><span class="dot" style="background:var(--red)"></span>Portfolio Heat Gauge</div></div>
    <canvas id="gaugeChart" height="160"></canvas><div style="text-align:center;margin-top:-10px"><span class="mono" style="font-size:28px;font-weight:900" id="gaugeVal"></span><br><small style="color:var(--muted)">Portfolio Heat %</small></div>
  </div>
  <div class="panel"><div class="panel-hdr"><div class="panel-title"><span class="dot" style="background:var(--purp)"></span>Sector Allocation</div></div>
    <canvas id="donutChart" height="160"></canvas>
  </div>
  <div class="panel"><div class="panel-hdr"><div class="panel-title"><span class="dot" style="background:var(--gold)"></span>Scenario Projection</div></div>
    <canvas id="scChart" height="160"></canvas>
  </div>
</div>

<div class="grid2">
  <div class="panel"><div class="panel-hdr"><div class="panel-title"><span class="dot" style="background:var(--green)"></span>P&amp;L per Stock</div></div>
    <canvas id="pnlChart" height="200"></canvas>
  </div>
  <div class="panel"><div class="panel-hdr"><div class="panel-title"><span class="dot" style="background:var(--blue)"></span>Price Levels</div></div>
    <canvas id="priceChart" height="200"></canvas>
  </div>
</div>

<div class="panel" style="margin-bottom:16px">
  <div class="panel-hdr">
    <div class="panel-title"><span class="dot" style="background:var(--gold)"></span>Stock Risk Table</div>
    <span class="panel-sub" id="tblCount"></span>
  </div>
  <div class="tbl-h tbl-wrap"><table id="riskTable"><thead><tr>
    <th>Stock</th><th>Sector</th><th>Entry</th><th>Current</th>
    <th>P&amp;L%</th><th>Dist SL%</th><th>Dist T1%</th><th>R:R</th><th>Risk</th><th>Action</th>
  </tr></thead><tbody id="tblBody"></tbody></table></div>
</div>

<div class="sc-grid" id="scCards"></div>

<div class="panel" style="margin-bottom:16px">
  <div class="panel-hdr"><div class="panel-title"><span class="dot" style="background:var(--purp)"></span>Sector Concentration</div></div>
  <div id="sectorBars"></div>
</div>

<div class="footer">⚠️ For information only — not financial advice. Capital: ₹{TOTAL_CAPITAL:,} · Run: {today}</div>
</div>

<script>
const STOCKS   = {stocks_js};
const KPIS     = {kpis_js};
const SCENARIOS= {sc_js};
const INVESTED = {invested};
const SEC_CLR  = {{"Oil & Gas":"#F85149","Defence":"#58A6FF","ETF":"#E3B341","Infra":"#BC8CFF","Energy":"#3FB950"}};
const FMT = v => '₹'+Math.abs(v).toLocaleString('en-IN',{{minimumFractionDigits:0,maximumFractionDigits:0}});
const PCT = v => (v>=0?'+':'')+((v*100).toFixed(2))+'%';

// ── Header
document.getElementById('updDate').textContent = STOCKS[0]?.Date || '';
document.getElementById('hdrSub').textContent =
  `${{STOCKS[0]?.Date||''}} · ${{STOCKS.length}} Stocks · ₹${{(KPIS['Total Invested']||0).toLocaleString('en-IN')}} Deployed`;

// ── Populate sector filter
const sectors = [...new Set(STOCKS.map(s=>s.Sector))];
const sf = document.getElementById('sectorFilter');
sectors.forEach(s=>{{ const o=document.createElement('option'); o.value=s; o.textContent=s; sf.appendChild(o); }});

// ── KPI cards
const kpiDefs = [
  {{label:'Total Capital',val:'₹'+((KPIS['Total Capital']||0).toLocaleString('en-IN')),sub:'Deployed: '+FMT(KPIS['Total Invested']||0),cls:'ck-blue',icon:'💰'}},
  {{label:'Current Value',val:FMT(KPIS['Total Current Value']||0),sub:'As of today',cls:'ck-blue',icon:'📈'}},
  {{label:'Total P&L',val:(KPIS['Total PnL Rs']>=0?'+':'')+FMT(KPIS['Total PnL Rs']||0),
    sub:PCT((KPIS['Total PnL Pct']||0)/100),cls:KPIS['Total PnL Rs']>=0?'ck-green':'ck-red',icon:'💹'}},
  {{label:'Portfolio Heat',val:(KPIS['Portfolio Heat Pct']||0).toFixed(2)+'%',sub:KPIS['Risk Status'],cls:'ck-gold',icon:'🛡️'}},
  {{label:'Capital After SL',val:FMT(KPIS['Capital After SL']||0),sub:'Worst-case scenario',cls:'ck-gold',icon:'🎯'}},
  {{label:'Active Stocks',val:KPIS['Active Stocks']||STOCKS.length,sub:sectors.length+' sectors',cls:'ck-purp',icon:'📊'}},
];
document.getElementById('kpiRow').innerHTML = kpiDefs.map(k=>
  `<div class="kpi ${{k.cls}}"><div style="font-size:20px;margin-bottom:8px">${{k.icon}}</div>
   <div class="kpi-label">${{k.label}}</div>
   <div class="kpi-val" style="color:${{k.cls.includes('green')?'var(--green)':k.cls.includes('red')?'var(--red)':k.cls.includes('gold')?'var(--gold)':k.cls.includes('purp')?'var(--purp)':'var(--blue)'}}">
   ${{k.val}}</div><div class="kpi-sub">${{k.sub}}</div></div>`).join('');

// ── Risk badge
const rb = document.getElementById('riskBadge');
const heat = KPIS['Portfolio Heat Pct']||0;
if(heat>=6){{rb.textContent='🛑 HARD STOP';rb.className='badge b-red';}}
else if(heat>=4){{rb.textContent='⚠ Warning Zone';rb.className='badge b-red';}}
else{{rb.textContent='✅ Safe Zone';rb.className='badge b-green';}}

// ── Gauge
const gClr = heat>=6?'#F85149':heat>=4?'#E3B341':'#3FB950';
document.getElementById('gaugeVal').textContent = heat.toFixed(2)+'%';
document.getElementById('gaugeVal').style.color = gClr;
new Chart(document.getElementById('gaugeChart'),{{
  type:'doughnut',
  data:{{datasets:[{{data:[heat,Math.max(0,10-heat)],
    backgroundColor:[gClr,'#1C2128'],borderColor:['transparent','transparent'],
    circumference:180,rotation:270,cutout:'75%',borderRadius:4}}]}},
  options:{{plugins:{{legend:{{display:false}},tooltip:{{enabled:false}}}},animation:{{duration:1200}}}}
}});

// ── Donut
const secMap={{}};
STOCKS.forEach(s=>{{secMap[s.Sector]=(secMap[s.Sector]||0)+(s['Investment Value']||0);}});
new Chart(document.getElementById('donutChart'),{{
  type:'doughnut',
  data:{{labels:Object.keys(secMap),datasets:[{{data:Object.values(secMap),
    backgroundColor:Object.keys(secMap).map(s=>SEC_CLR[s]||'#8B949E'),
    borderColor:'#0D1117',borderWidth:3,hoverOffset:6}}]}},
  options:{{cutout:'62%',plugins:{{legend:{{position:'right',labels:{{color:'#8B949E',font:{{size:10}},padding:6,boxWidth:8}}}},
    tooltip:{{callbacks:{{label:c=>`${{c.label}}: ${{FMT(c.parsed)}}`}}}}}}}}
}});

// ── Scenario chart
new Chart(document.getElementById('scChart'),{{
  type:'bar',
  data:{{labels:SCENARIOS.map(s=>s.label),
    datasets:[{{label:'Projected',data:SCENARIOS.map(s=>s.value),
      backgroundColor:['rgba(63,185,80,.6)','rgba(88,166,255,.6)','rgba(248,81,73,.6)'],
      borderColor:['#3FB950','#58A6FF','#F85149'],borderWidth:2,borderRadius:5}},
      {{label:'Invested',data:SCENARIOS.map(()=>INVESTED),
        backgroundColor:'rgba(139,148,158,.15)',borderColor:'#8B949E',borderWidth:1}}]}},
  options:{{plugins:{{legend:{{labels:{{color:'#8B949E',font:{{size:10}}}}}}}},
    scales:{{x:{{ticks:{{color:'#8B949E',font:{{size:9}}}},grid:{{color:'rgba(48,54,61,.4)'}}}},
             y:{{ticks:{{color:'#8B949E',font:{{size:9}},callback:v=>'₹'+Math.round(v/1000)+'K'}},grid:{{color:'rgba(48,54,61,.4)'}}}}}}}}
}});

// ── P&L Chart
const pnlVals = STOCKS.map(s=>+((s.PnL_Pct||s['PnL Pct']||0)*100).toFixed(2));
new Chart(document.getElementById('pnlChart'),{{
  type:'bar',
  data:{{labels:STOCKS.map(s=>s.Stock.substring(0,10)),
    datasets:[{{label:'P&L%',data:pnlVals,
      backgroundColor:pnlVals.map(v=>v>=0?'rgba(63,185,80,.7)':'rgba(248,81,73,.7)'),
      borderColor:pnlVals.map(v=>v>=0?'#3FB950':'#F85149'),borderWidth:1.5,borderRadius:4}}]}},
  options:{{plugins:{{legend:{{display:false}}}},
    scales:{{x:{{ticks:{{color:'#8B949E',font:{{size:9}}}},grid:{{color:'rgba(48,54,61,.4)'}}}},
             y:{{ticks:{{color:'#8B949E',font:{{size:9}},callback:v=>v+'%'}},grid:{{color:'rgba(48,54,61,.4)'}}}}}}}}
}});

// ── Price levels
const nm = (base,v)=>base?+((( v-base)/base*100).toFixed(2)):0;
new Chart(document.getElementById('priceChart'),{{
  type:'bar',
  data:{{labels:STOCKS.map(s=>s.Stock.substring(0,8)),
    datasets:[
      {{label:'Current%',data:STOCKS.map(s=>nm(s['Entry Price'],s['Current Price'])),backgroundColor:'rgba(88,166,255,.6)',borderColor:'#58A6FF',borderWidth:1.5,borderRadius:3}},
      {{label:'SL%',data:STOCKS.map(s=>nm(s['Entry Price'],s['Stop Loss'])),backgroundColor:'rgba(248,81,73,.5)',borderColor:'#F85149',borderWidth:1.5,borderRadius:3}},
      {{label:'T1%',data:STOCKS.map(s=>nm(s['Entry Price'],s['Target 1'])),backgroundColor:'rgba(63,185,80,.5)',borderColor:'#3FB950',borderWidth:1.5,borderRadius:3}},
    ]}},
  options:{{plugins:{{legend:{{labels:{{color:'#8B949E',font:{{size:10}}}}}}}},
    scales:{{x:{{ticks:{{color:'#8B949E',font:{{size:9}}}},grid:{{color:'rgba(48,54,61,.4)'}}}},
             y:{{ticks:{{color:'#8B949E',font:{{size:9}},callback:v=>v+'%'}},grid:{{color:'rgba(48,54,61,.4)'}}}}}}}}
}});

// ── Table
function buildTable(data){{
  document.getElementById('tblCount').textContent='Showing '+data.length+' stocks';
  const tbody=document.getElementById('tblBody'); tbody.innerHTML='';
  data.forEach(r=>{{
    const pnl=r.PnL_Pct||r['PnL Pct']||0;
    const pCls=pnl>=0?'pos':'neg';
    const sl=r.Dist_to_SL_Pct||r['Dist to SL Pct']||0;
    const slCls=sl<0?'neg':sl<0.03?'warn':'pos';
    const rr=r.Reward_Risk||r['Reward Risk']||0;
    const ac=r.Action||'Hold';
    const acCls=ac==='EXIT NOW'?'ab-exit':ac==='Hold'?'ab-hold':ac.startsWith('Reduce')?'ab-reduce':'ab-book';
    const tr=document.createElement('tr');
    tr.innerHTML=`<td>${{r.Stock}}</td>
      <td><span style="font-size:10px;padding:2px 7px;border-radius:10px;background:rgba(0,0,0,.3);color:var(--muted);border:1px solid var(--b)">${{r.Sector}}</span></td>
      <td class="mono">₹${{(r['Entry Price']||0).toFixed(2)}}</td>
      <td class="mono">₹${{(r['Current Price']||0).toFixed(2)}}</td>
      <td class="mono ${{pCls}}">${{PCT(pnl)}}</td>
      <td class="mono ${{slCls}}">${{(sl*100).toFixed(2)}}%</td>
      <td class="mono" style="color:var(--blue)">${{((r.Dist_to_T1_Pct||r['Dist to T1 Pct']||0)*100).toFixed(2)}}%</td>
      <td class="mono ${{rr>=1.5?'pos':rr>=1?'warn':'neg'}}">${{rr.toFixed(2)}}x</td>
      <td class="mono" style="color:var(--red)">₹${{(r['Risk Amount Rs']||0).toFixed(0)}}</td>
      <td><span class="ab ${{acCls}}">${{ac}}</span></td>`;
    tbody.appendChild(tr);
  }});
}}
buildTable(STOCKS);
function filterSector(sec){{ buildTable(sec==='all'?STOCKS:STOCKS.filter(s=>s.Sector===sec)); }}

// ── Scenario cards
document.getElementById('scCards').innerHTML=SCENARIOS.map((sc,i)=>{{
  const clr=i===0?'var(--green)':i===1?'var(--blue)':'var(--red)';
  const em=i===0?'🟢':i===1?'🔵':'🔴';
  return `<div class="sc" style="border-color:${{clr}}40">
    <div style="font-size:26px;margin-bottom:8px">${{em}}</div>
    <div class="sc-name" style="color:${{clr}}">${{sc.label.toUpperCase()}}</div>
    <div class="sc-val" style="color:${{clr}}">₹${{(sc.value||0).toLocaleString('en-IN',{{maximumFractionDigits:0}})}}</div>
    <div class="sc-prob">${{Math.round((sc.prob||0)*100)}}% Probability</div>
  </div>`;
}}).join('');

// ── Sector bars
const sbDiv=document.getElementById('sectorBars');
const secTotals=Object.entries(secMap).sort((a,b)=>b[1]-a[1]);
const totInv=KPIS['Total Invested']||1;
sbDiv.innerHTML=secTotals.map(([sec,val])=>{{
  const p=(val/totInv*100).toFixed(1);
  const clr=SEC_CLR[sec]||'#8B949E';
  const warn=parseFloat(p)>40;
  return `<div class="sb-row">
    <span class="sb-label" style="color:${{warn?'var(--red)':'var(--muted)'}}">${{sec}}${{warn?' ⚠':''}}</span>
    <div class="sb-track"><div class="sb-fill" style="width:${{Math.min(p,100)}}%;background:${{clr}}"></div></div>
    <span class="sb-pct ${{warn?'neg':''}}">${{p}}%</span>
  </div>`;
}}).join('');
</script>
</body>
</html>"""
    tmp_path = out_path + ".tmp"
    with open(tmp_path, "w", encoding="utf-8") as f:
        f.write(html)
    
    # Scenario 8 — Atomic replace (browser never sees a partially written file)
    if os.path.exists(out_path):
        os.remove(out_path)
    os.rename(tmp_path, out_path)
    log.info("HTML dashboard saved: %s", os.path.basename(out_path))

def export_status_csv(kpis: dict, out_path: str):
    """Scenario 7 — Export a small CSV with the last run timestamp for Power BI alerts."""
    now = datetime.datetime.now()
    status_data = [
        {"Metric": "Last Pipeline Run", "Value": now.strftime("%Y-%m-%d %H:%M:%S")},
        {"Metric": "Pipeline Status", "Value": "Success"},
        {"Metric": "Active Stocks", "Value": kpis.get("Active Stocks", 0)},
        {"Metric": "Portfolio Heat", "Value": f"{kpis.get('Portfolio Heat Pct', 0):.2f}%"},
    ]
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["Metric", "Value"])
        w.writeheader()
        w.writerows(status_data)
    log.info("Pipeline Status CSV saved: %s", os.path.basename(out_path))

# ══════════════════════════════════════════════════════════════════════════
# 6. MAIN PIPELINE
# ══════════════════════════════════════════════════════════════════════════
def main():
    start_time = time.time()
    log.info("="*60)
    log.info("Tactical Trading Analytics — Build Pipeline")
    log.info("="*60)
    
    # ── 0. Pre-flight & Backup (Scenarios 9, 10)
    preflight(SOURCE_XL)
    backup_source(SOURCE_XL)
    
    today = datetime.date.today().isoformat()

    # ── 1. Load (Scenario 2, 3, 5)
    raw_records = load_portfolio(SOURCE_XL)
    scenarios   = load_scenarios(SOURCE_XL)
    price_rows  = load_prices(SOURCE_XL)

    # ── 2. Compute (Scenario 1, 4, 6)
    rows = [compute_row(r, today) for r in raw_records]
    kpis = compute_kpis(rows, TOTAL_CAPITAL)

    log.info("")
    log.info("── Portfolio Snapshot ──────────────")
    log.info("  Stocks   : %d", len(rows))
    log.info("  Invested : ₹%s", f"{kpis['Total Invested']:,.2f}")
    log.info("  Value    : ₹%s", f"{kpis['Total Current Value']:,.2f}")
    log.info("  P&L      : ₹%s (%.2f%%)", f"{kpis['Total PnL Rs']:+,.2f}", kpis['Total PnL Pct'])
    log.info("  Heat     : %.2f%% [%s]", kpis['Portfolio Heat Pct'], kpis['Risk Status'])
    log.info("")

    # Action summary
    from collections import Counter
    actions = Counter(r["Action"] for r in rows)
    for ac, cnt in actions.items():
        flag = " ← ACTION NEEDED" if ac != "Hold" else ""
        log.info("  Action [%s]: %d stock(s)%s", ac, cnt, flag)

    log.info("")

    # ── 3. Write outputs (Scenario 7, 8)
    log.info("Writing outputs …")
    write_excel(rows, kpis, scenarios, price_rows, OUT_XL)
    export_csvs(rows, scenarios, price_rows, OUT_DIR)
    export_status_csv(kpis, STATUS_CSV)
    generate_html(rows, kpis, scenarios, OUT_HTML)

    # ── 4. Write computed values back to SOURCE Excel (NEW)
    write_back_source(rows, SOURCE_XL)

    # ── Verify Formulas (Scenario 11)
    _verify_formulas(OUT_XL)

    elapsed = time.time() - start_time
    log.info("")
    log.info("="*60)
    log.info("BUILD COMPLETE in %.2f seconds", elapsed)
    log.info("  Excel   : %s", os.path.basename(OUT_XL))
    log.info("  HTML    : %s", os.path.basename(OUT_HTML))
    log.info("  Status  : %s", os.path.basename(STATUS_CSV))
    log.info("  CSVs    : Fact_Portfolio | Dim_Stock | Dim_Scenario | Fact_Prices")
    log.info("  Log     : build_model.log")
    log.info("="*60)
    return rows, kpis

def write_back_source(rows: list[dict], source_path: str):
    """
    Write all computed values back to the SOURCE Excel (Tactical_Model_INSTITUTIONAL.xlsx)
    Fact_Portfolio sheet so it stays in sync after every build_model.py run.

    Column mapping in source (confirmed from header analysis):
      G=Target 1, H=Target 2, I=Target 3, J=Stop Loss, K=Risk Level
      L=Investment Value, N=Current Value, O=P&L Rs, P=P&L %,
      Q=Risk Amount, R=SL Exit, S=Dist SL, T=Dist T1,
      U=RR Ratio, V=Exp Value, W=Action,
      X=Strategic Role, Y=Volatility
    """
    from openpyxl.utils import column_index_from_string as _ci
    try:
        wb = openpyxl.load_workbook(source_path)
        ws = wb["Fact_Portfolio"]

        # Build a map: stock_name → row_index (searching from row 3 onwards)
        stock_row_map = {}
        for ri in range(3, ws.max_row + 1):
            val = ws.cell(ri, 1).value
            if val:
                stock_row_map[str(val).strip()] = ri

        WRITE_MAP = {
            # col_letter: computed_key
            "D":  "Allocation Rs",  # SYNC: Allocation Rs now updates Column D
            "G":  "Target 1",
            "H":  "Target 2",
            "I":  "Target 3",
            "J":  "Stop Loss",
            "K":  "Risk Level",
            "L":  "Investment Value",
            "N":  "Current Value",
            "O":  "PnL Rs",
            "P":  "PnL Pct",
            "Q":  "Risk Amount Rs",
            "R":  "SL Exit Value",
            "S":  "Dist to SL Pct",
            "T":  "Dist to T1 Pct",
            "U":  "Reward Risk",
            "V":  "Expected Value",
            "W":  "Action",
            "X":  "Strategic Role",
            "Y":  "Volatility",
        }

        NUMBER_FMTS = {
            "D": "#,##0.00", # Allocation Rs
            "G": "#,##0.00", "H": "#,##0.00", "I": "#,##0.00",
            "J": "#,##0.00", "L": "#,##0.00", "N": "#,##0.00",
            "O": '#,##0.00;[Red]-#,##0.00',
            "P": "0.00%",
            "Q": "#,##0.00", "R": "#,##0.00",
            "S": "0.00%", "T": "0.00%",
            "U": "0.00",   "V": "#,##0.00",
        }

        updated = 0
        for row in rows:
            stock = str(row.get("Stock", "")).strip()
            if stock not in stock_row_map:
                continue
            ri = stock_row_map[stock]
            for col_letter, key in WRITE_MAP.items():
                val = row.get(key)
                if val is None:
                    continue
                cell = ws.cell(ri, _ci(col_letter))
                cell.value = val
                if col_letter in NUMBER_FMTS:
                    cell.number_format = NUMBER_FMTS[col_letter]
            updated += 1

        try:
            wb.save(source_path)
            log.info("Source write-back: %d stocks updated in %s",
                     updated, os.path.basename(source_path))
        except PermissionError:
            log.warning("Source write-back SKIPPED — %s is open in Excel. Close it and re-run.",
                        os.path.basename(source_path))
    except Exception as e:
        log.warning("Source write-back failed: %s", e)


def _verify_formulas(path):
    """Small check to ensure formulas were actually written."""
    wb = openpyxl.load_workbook(path, data_only=False) # formulas preserved
    ws = wb["Portfolio_Master"]
    f_val = ws["G2"].value # Column G is Target 1 in new write (7th col)
    # Wait, check mapping: FP_HEADERS[6] = Target 1. 0-based index 6 is Column G.
    if f_val and str(f_val).startswith("="):
        log.info("Formula verification: Portfolio_Master G2 has formula %s", f_val)
    else:
        log.warning("Formula check FAILED: Portfolio_Master G2 has value %s", f_val)
    
    ws2 = wb["Stock_Visualization_Data"]
    g_val = ws2["G2"].value
    if g_val and str(g_val).startswith("="):
        log.info("Formula verification: Stock_Visualization_Data G2 has formula %s", g_val)
    else:
        log.warning("Formula check FAILED: Stock_Visualization_Data G2 has value %s", g_val)

if __name__ == "__main__":
    main()
