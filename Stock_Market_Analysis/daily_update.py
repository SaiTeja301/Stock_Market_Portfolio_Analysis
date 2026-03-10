"""
daily_update.py  —  One-command Daily Price Update
===================================================
1. Reads current stocks from Tactical_Model_INSTITUTIONAL.xlsx
2. Shows today's last price for each stock
3. Lets you type the new current price (press Enter to keep last)
4. Updates the Excel file in place
5. Runs build_model.py  →  regenerates all CSVs + HTML dashboard
6. Opens tactical_dashboard.html in your browser

Usage:  python daily_update.py
"""

import os, sys, webbrowser, subprocess, datetime
import openpyxl

BASE     = r"E:\Stock Market\Analysis\Prompts\MAR_WEEK_1 Plan"
XL_PATH  = os.path.join(BASE, "Tactical_Model_INSTITUTIONAL.xlsx")
HTML_PATH= os.path.join(BASE, "tactical_dashboard.html")
SCRIPT   = os.path.join(BASE, "build_model.py")

# ── ANSI colours (Windows 10+ supports these)
GRN="\033[92m"; YLW="\033[93m"; RED="\033[91m"
BLU="\033[94m"; CYN="\033[96m"; RST="\033[0m"; BLD="\033[1m"

def clr(text, c): return f"{c}{text}{RST}"
def enable_ansi():
    if sys.platform == "win32":
        import ctypes
        k = ctypes.windll.kernel32
        k.SetConsoleMode(k.GetStdHandle(-11), 7)

def banner():
    print()
    print(clr("═"*62, BLU))
    print(clr("  📊  TACTICAL TRADING — DAILY PRICE UPDATE", BLD))
    print(clr(f"  {datetime.date.today().isoformat()}  ·  Power BI Pipeline", CYN))
    print(clr("═"*62, BLU))
    print()

def load_fact_portfolio(path):
    """Return (wb, ws, headers, data_rows) from Fact_Portfolio sheet."""
    if not os.path.exists(path):
        print(clr(f"❌  File not found: {path}", RED))
        sys.exit(1)
    wb = openpyxl.load_workbook(path)

    # Find the right sheet
    sheet = "Fact_Portfolio" if "Fact_Portfolio" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet]

    rows = list(ws.iter_rows(values_only=False))  # keep cell references
    if not rows:
        print(clr("❌  Fact_Portfolio sheet is empty.", RED))
        sys.exit(1)

    # Header row
    headers = [c.value for c in rows[0]]
    return wb, ws, headers, rows[1:]   # (workbook, worksheet, header list, data rows)

def find_col(headers, *candidates):
    """Find column index (0-based) by trying candidate names."""
    for name in candidates:
        for i, h in enumerate(headers):
            if h and str(h).strip().lower() == name.lower():
                return i
    return None

def get_float(prompt, default):
    """Prompt user for a float; blank = keep default."""
    try:
        raw = input(prompt).strip()
        if raw == "":
            return default, False
        v = float(raw)
        return v, True
    except ValueError:
        print(clr("  ⚠  Invalid — keeping previous value", YLW))
        return default, False

def main():
    enable_ansi()
    banner()

    # Load workbook
    print(clr("Loading portfolio…", CYN))
    wb, ws, headers, data_rows = load_fact_portfolio(XL_PATH)

    # Find key columns
    col_stock   = find_col(headers, "Stock", "Stock Name")
    col_entry   = find_col(headers, "Entry Price", "Entry")
    col_current = find_col(headers, "Current Price", "Current")
    col_sl      = find_col(headers, "Stop Loss")
    col_t1      = find_col(headers, "Target 1")
    col_pnl_pct = find_col(headers, "PnL Pct", "P&L %", "P&L%")

    if col_stock is None or col_current is None:
        print(clr("❌  Could not find 'Stock' or 'Current Price' columns.", RED))
        print(clr(f"   Available columns: {headers}", YLW))
        sys.exit(1)

    # Filter non-empty data rows
    stocks = []
    for row in data_rows:
        val = row[col_stock].value
        if val:
            stocks.append(row)

    if not stocks:
        print(clr("❌  No stock data found.", RED))
        sys.exit(1)

    print(clr(f"  Found {len(stocks)} stocks\n", GRN))
    print(f"{'#':<4} {'Stock':<28} {'Entry':>10} {'Last':>10} {'SL':>10} {'T1':>10}")
    print("─"*68)
    for i, row in enumerate(stocks, 1):
        stk  = str(row[col_stock].value)
        ent  = row[col_entry].value  if col_entry  is not None else 0
        cur  = row[col_current].value if col_current is not None else row[col_entry].value
        sl   = row[col_sl].value     if col_sl    is not None else 0
        t1   = row[col_t1].value     if col_t1    is not None else 0
        print(f"{i:<4} {stk:<28} {(ent or 0):>10.2f} "
              f"{clr(f'{(cur or 0):>10.2f}', CYN)}"
              f" {(sl or 0):>10.2f} {(t1 or 0):>10.2f}")

    print()
    print(clr("─"*68, BLU))
    print(clr("Enter today's prices below.", BLD))
    print(clr("Press Enter to keep the previous value. Type 'q' to quit.\n", YLW))

    # Collect new prices
    updates = {}
    for i, row in enumerate(stocks, 1):
        stk = str(row[col_stock].value)
        old = row[col_current].value or row[col_entry].value or 0
        prompt = f"  [{i:>2}] {stk:<28} (last ₹{old:.2f}) → ₹"
        new_val, changed = get_float(prompt, old)
        if new_val == "q":
            print(clr("\n  Aborted — no changes saved.", YLW))
            sys.exit(0)
        updates[i-1] = (row, new_val, changed)

    # Summary
    print()
    print(clr("─"*68, BLU))
    print(clr("  CHANGES SUMMARY:", BLD))
    any_change = False
    for idx, (row, new_val, changed) in updates.items():
        stk = str(row[col_stock].value)
        old = row[col_current].value or row[col_entry].value or 0
        if changed:
            diff = new_val - old
            sym  = "▲" if diff >= 0 else "▼"
            clr_ = GRN if diff >= 0 else RED
            print(f"  {stk:<28}  {old:.2f} → "
                  f"{clr(f'{new_val:.2f}  {sym}{abs(diff):.2f}', clr_)}")
            any_change = True

    if not any_change:
        print(clr("  No prices changed.", YLW))
        go = input(clr("\n  Run pipeline anyway? [y/N] ", BLD)).strip().lower()
        if go != "y":
            print(clr("  Exiting — nothing updated.", YLW))
            sys.exit(0)

    # Confirm
    print()
    confirm = input(clr("  Save changes and run pipeline? [Y/n] ", BLD)).strip().lower()
    if confirm == "n":
        print(clr("  Aborted — no changes saved.", YLW))
        sys.exit(0)

    # ── Write back to Excel
    print()
    print(clr("  Saving prices to Excel…", CYN))

    # Also find the date column to update
    col_date = find_col(headers, "Date")

    for idx, (row, new_val, changed) in updates.items():
        if col_current is not None:
            # Row is tuple of cells; column index is 0-based, openpyxl row index 1-based
            actual_row = row[col_stock].row
            ws.cell(row=actual_row, column=col_current+1, value=new_val)
        if col_date is not None:
            ws.cell(row=row[col_stock].row, column=col_date+1,
                    value=datetime.date.today().isoformat())

    wb.save(XL_PATH)
    print(clr(f"  ✅  Saved: {os.path.basename(XL_PATH)}", GRN))

    # ── Run build_model.py
    print()
    print(clr("  Running build_model.py…", CYN))
    print(clr("─"*68, BLU))
    result = subprocess.run([sys.executable, SCRIPT], cwd=BASE)

    if result.returncode != 0:
        print(clr("\n  ❌  build_model.py failed — check errors above.", RED))
        sys.exit(1)

    print(clr("─"*68, BLU))
    print(clr("  ✅  Pipeline complete!", GRN))
    print()

    # ── Open browser
    print(clr("  Opening dashboard in browser…", CYN))
    webbrowser.open(f"file:///{HTML_PATH.replace(os.sep, '/')}")

    print()
    print(clr("═"*62, GRN))
    print(clr("  DONE — Refresh Power BI to see updated data", BLD))
    print(clr("═"*62, GRN))
    print()
    print(clr("  In Power BI:", YLW))
    print("    Home → Refresh  (one click — all charts update)")
    print()

if __name__ == "__main__":
    main()
